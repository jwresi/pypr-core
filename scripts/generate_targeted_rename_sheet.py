from __future__ import annotations
import csv
import re
from collections import defaultdict, Counter
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from packages.jake.connectors.mcp.jake_ops_mcp import JakeOps, DEVICE_LABEL_RE

OUTDIR = Path('output/spreadsheet')
OUTDIR.mkdir(parents=True, exist_ok=True)
CSV_PATH = OUTDIR / 'netbox_targeted_rename_proposals.csv'
XLSX_PATH = OUTDIR / 'netbox_targeted_rename_proposals.xlsx'

TARGETS = [
    (136, '104 Tapscott V5000'), (48, '170 Tapscott V5000'), (137, '175 Tapscott V1000'),
    (221, '225 Buffalo V5000'), (171, '324 Howard Ave V2000'), (176, '324 Howard Ave V5000'),
    (63, '508 Howard V5000'), (138, '511 Howard V3000'), (139, '511 Howard V5000'),
    (64, '545 Ralph Ave V1000'), (66, '595 Howard Ave V1000'), (140, '610 Howard V3000'),
    (141, '610 Howard V5000'), (142, '692 ralph ave V3000'), (249, '721 Fenimore V1000'),
    (70, '725 Howard - 1145 Lenox'), (143, '725 Howard V5000'), (67, '728 E NY - 955 Rutland'),
    (250, '728 E New York V5000'), (144, '728 Howard Ave V1000'), (68, '955 Rutland - 728 E NY'),
    (35, '955 Rutland - 1145 Lenox'), (145, '1142 Lenox Rd V1000'), (146, '1144 Lenox Ave V1000'),
    (69, '1145 Lenox - 725 Howard'), (36, '1145 Lenox - 955 Rutland'), (147, '1145 Lenox Ave V5000'),
    (162, '1196 E NY Ave V2000'), (182, '1371 St Marks Ave V2000'), (172, '1371 St Marks Ave V5000'),
    (175, '1629 Park Pl V1000'), (169, '1640 Park Pl V3000'), (156, '1640 Park Pl V5000'),
    (163, '1640 Sterling Pl V2000'), (148, '1691 St Johns Pl V5000'), (149, '1724 Sterling Pl V5000'),
    (150, '1766 Sterling Pl V1000'), (62, '1767 Sterling Pl V5000'), (65, '1790 Sterling Pl V1000'),
    (177, '2041 Pacific St V1000'), (151, '2045 Union V5000'), (170, '2058 Union V3000'),
    (152, '2058 union V5000'), (153, '2069 Union V5000'),
    (128, 'Cambridge Square Building 1 V5000'), (129, 'Cambridge Square Building 2 V2000'),
    (130, 'Cambridge Square Building 3 V2000'), (131, 'Cambridge Square Building 4 V2000'),
    (132, 'Cambridge Square Building 5 V5000'), (133, 'Cambridge Square Building 6 V2000'),
    (134, 'Cambridge Square Building 7 V2000'), (135, 'Cambridge Square Clubhouse V2000'),
    (320, 'Savoy - Building 7 v5000'), (319, 'Savoy Building 1 v5000'), (321, 'Savoy Building 2 v5000'),
    (322, 'Savoy Building 3 v5000'), (323, 'Savoy Building 4 v5000'), (324, 'Savoy Building 5 v5000'),
    (325, 'Savoy Building 6 v5000'),
]

CAMBRIDGE_OVERRIDES = {
    'cambridge bldg1': '000004.001',
    'cambridge square building 1': '000004.001',
    'cambridge bldg2': '000004.002',
    'cambridge square building 2': '000004.002',
    'cambridge bldg3': '000004.003',
    'cambridge square building 3': '000004.003',
    'cambridge bldg4': '000004.004',
    'cambridge square building 4': '000004.004',
    'cambridge bldg5': '000004.005',
    'cambridge square building 5': '000004.005',
    'cambridge bldg6': '000004.006',
    'cambridge square building 6': '000004.006',
    'cambridge bldg7': '000004.007',
    'cambridge square building 7': '000004.007',
    'cambridge office': '000004.008',
    'cambridge square clubhouse': '000004.008',
}
SAVOY_OVERRIDES = {
    'savoy bldg1': '000002.004', 'savoy building 1': '000002.004',
    'savoy bldg2': '000002.005', 'savoy building 2': '000002.005',
    'savoy bldg3': '000002.006', 'savoy building 3': '000002.006',
    'savoy bldg4': '000002.007', 'savoy building 4': '000002.007',
    'savoy bldg5': '000002.008', 'savoy building 5': '000002.008',
    'savoy bldg6': '000002.009', 'savoy building 6': '000002.009',
    'savoy bldg7': '000002.010', 'savoy building 7': '000002.010',
}
COLOR_BY_CONF = {'high': 'E2F0D9', 'medium': 'FFF2CC', 'low': 'FCE4D6', 'none': 'F4CCCC'}


def normalize_text(s: str) -> str:
    return re.sub(r'\s+', ' ', (s or '').strip().lower())


def canonical_site(site_obj):
    return str((site_obj or {}).get('slug') or (site_obj or {}).get('name') or '').strip()


def location_name(dev):
    return str((dev.get('location') or {}).get('display') or (dev.get('location') or {}).get('name') or '').strip()


def location_desc(dev):
    return str((dev.get('location') or {}).get('description') or '').strip()


def family_code(model: str, current_name: str) -> str:
    s = f'{model} {current_name}'.lower()
    if 'v5000' in s:
        return 'V5K'
    if 'v3000' in s:
        return 'V3K'
    if 'v2000' in s:
        return 'V2K'
    if 'v1000' in s:
        return 'V1K'
    if 'eh-8010fx' in s:
        return 'EH'
    return 'DEV'


def override_prefix(current_name: str, loc_desc: str) -> tuple[str, str, str]:
    hay = f"{normalize_text(current_name)} | {normalize_text(loc_desc)}"
    for k, prefix in SAVOY_OVERRIDES.items():
        if k in hay:
            return prefix, 'medium', 'Savoy building override extrapolated from user-provided Building 1 example'
    for k, prefix in CAMBRIDGE_OVERRIDES.items():
        if k in hay:
            return prefix, 'low', 'Cambridge building override inferred from building number/clubhouse naming'
    return '', 'none', ''


def main() -> None:
    ops = JakeOps()
    all_devices = ops._netbox_all_devices()
    by_id = {d['id']: d for d in all_devices}

    valid_prefixes_by_location = defaultdict(set)
    for d in all_devices:
        name = str(d.get('name') or '').strip()
        if not DEVICE_LABEL_RE.match(name):
            continue
        loc = location_name(d)
        m = re.match(r'^(\d{6}\.\d{3})\.', name)
        if loc and m:
            valid_prefixes_by_location[loc].add(m.group(1))

    rows = []
    for target_id, _ in TARGETS:
        d = by_id[target_id]
        current_name = str(d.get('name') or '')
        site_code = canonical_site(d.get('site'))
        loc = location_name(d)
        loc_desc = location_desc(d)
        model = str((d.get('device_type') or {}).get('model') or '')
        role = str((d.get('role') or {}).get('name') or '')
        fam = family_code(model, current_name)
        exact_prefixes = sorted(valid_prefixes_by_location.get(loc, set()))
        proposed_prefix, confidence, basis = override_prefix(current_name, loc_desc)
        notes = ''
        if not proposed_prefix:
            if len(exact_prefixes) == 1:
                proposed_prefix = exact_prefixes[0]
                confidence = 'high'
                basis = 'Exact same-location valid NetBox device already uses this prefix'
            elif len(exact_prefixes) > 1:
                site_matched = [p for p in exact_prefixes if p.startswith(f'{site_code}.')]
                if len(site_matched) == 1:
                    proposed_prefix = site_matched[0]
                    confidence = 'medium'
                    basis = 'Multiple same-location prefixes exist; selected the only prefix matching the device site'
                    notes = f'all same-location prefixes={", ".join(exact_prefixes)}'
                else:
                    confidence = 'none'
                    basis = 'Could not safely choose among multiple same-location valid prefixes'
                    notes = f'ambiguous same-location prefixes={", ".join(exact_prefixes)}'
            else:
                confidence = 'none'
                basis = 'No exact same-location valid prefix and no explicit building override found'
        rows.append({
            'netbox_id': target_id,
            'current_name': current_name,
            'site_code': site_code,
            'location': loc,
            'location_description': loc_desc,
            'model': model,
            'role': role,
            'exact_location_valid_prefixes': ', '.join(exact_prefixes),
            'proposed_prefix': proposed_prefix,
            'device_family_code': fam,
            'sequence': '',
            'proposed_name': '',
            'confidence': confidence,
            'basis': basis,
            'notes': notes,
        })

    seq_counters = defaultdict(int)
    for row in sorted(rows, key=lambda r: (r['proposed_prefix'], r['device_family_code'], r['current_name'])):
        if not row['proposed_prefix']:
            continue
        key = (row['proposed_prefix'], row['device_family_code'])
        seq_counters[key] += 1
        row['sequence'] = f"{seq_counters[key]:02d}"
        row['proposed_name'] = f"{row['proposed_prefix']}.{row['device_family_code']}{row['sequence']}"

    fieldnames = [
        'netbox_id', 'current_name', 'site_code', 'location', 'location_description', 'model', 'role',
        'exact_location_valid_prefixes', 'proposed_prefix', 'device_family_code', 'sequence',
        'proposed_name', 'confidence', 'basis', 'notes'
    ]

    with CSV_PATH.open('w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Rename Proposals'
    ws.freeze_panes = 'A2'
    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(color='FFFFFF', bold=True)
    for col, name in enumerate(fieldnames, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for r_idx, row in enumerate(rows, start=2):
        fill = PatternFill('solid', fgColor=COLOR_BY_CONF[row['confidence']])
        for c_idx, name in enumerate(fieldnames, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=row[name])
            cell.fill = fill
            if name in ('basis', 'notes', 'location', 'location_description'):
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    widths = {'A':10,'B':32,'C':10,'D':38,'E':28,'F':28,'G':12,'H':24,'I':14,'J':12,'K':10,'L':22,'M':12,'N':56,'O':36}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    summary = wb.create_sheet('Summary')
    summary.append(['metric', 'value'])
    summary['A1'].fill = header_fill; summary['B1'].fill = header_fill
    summary['A1'].font = header_font; summary['B1'].font = header_font
    counts = Counter(r['confidence'] for r in rows)
    for item in [
        ('total_rows', len(rows)),
        ('high_confidence', counts['high']),
        ('medium_confidence', counts['medium']),
        ('low_confidence', counts['low']),
        ('unresolved', counts['none']),
        ('with_proposed_name', sum(1 for r in rows if r['proposed_name'])),
    ]:
        summary.append(item)
    summary.column_dimensions['A'].width = 24
    summary.column_dimensions['B'].width = 14
    wb.save(XLSX_PATH)
    print(CSV_PATH)
    print(XLSX_PATH)

if __name__ == '__main__':
    main()
